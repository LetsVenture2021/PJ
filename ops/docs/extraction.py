"""Bounded, sanitizing preview extraction for uploaded documents.

Extractors convert accepted uploads into Markdown previews for the assistant.
They never execute uploaded content: notebooks are read as JSON with outputs
stripped, HTML is reduced to text through an allowlist parser, SVG is
structurally inspected with unsafe nodes counted rather than rendered, and
header-only formats surface metadata without materializing tensors. All
extracted text passes through secret redaction as defense in depth.
"""

from __future__ import annotations

import csv
import json
import re
import struct
from html.parser import HTMLParser
from pathlib import Path
from xml.etree import ElementTree

from ops.docs.formats import Classification, Family


class ExtractionError(RuntimeError):
    """Raised when a preview cannot be produced; the artifact stays registered."""

    def __init__(self, code: str):
        super().__init__(code)
        self.code = code


MAX_TABULAR_SAMPLE_ROWS = 50
MAX_JSONL_SAMPLE_RECORDS = 20

LANGUAGE_BY_EXTENSION = {
    ".py": "python",
    ".pyi": "python",
    ".js": "javascript",
    ".mjs": "javascript",
    ".cjs": "javascript",
    ".ts": "typescript",
    ".tsx": "tsx",
    ".jsx": "jsx",
    ".go": "go",
    ".rs": "rust",
    ".java": "java",
    ".kt": "kotlin",
    ".swift": "swift",
    ".c": "c",
    ".h": "c",
    ".cc": "cpp",
    ".cpp": "cpp",
    ".cxx": "cpp",
    ".hpp": "cpp",
    ".cs": "csharp",
    ".rb": "ruby",
    ".php": "php",
    ".sh": "bash",
    ".bash": "bash",
    ".zsh": "bash",
    ".ps1": "powershell",
    ".sql": "sql",
    ".r": "r",
    ".jl": "julia",
    ".lua": "lua",
    ".proto": "protobuf",
    ".graphql": "graphql",
    ".gql": "graphql",
    ".tf": "hcl",
    ".hcl": "hcl",
    ".yaml": "yaml",
    ".yml": "yaml",
    ".toml": "toml",
    ".json": "json",
    ".scala": "scala",
    ".dart": "dart",
    ".zig": "zig",
}

_SECRET_PATTERNS = tuple(
    re.compile(pattern)
    for pattern in (
        r"sk-[A-Za-z0-9_\-]{20,}",
        r"gh[pousr]_[A-Za-z0-9]{20,}",
        r"AKIA[0-9A-Z]{16}",
        r"xox[baprs]-[A-Za-z0-9\-]{10,}",
        r"-----BEGIN [A-Z ]*PRIVATE KEY-----",
        r"(?i)(api[_-]?key|secret|password|token)\s*[:=]\s*['\"][^'\"]{8,}['\"]",
    )
)


def redact(text: str) -> str:
    """Uploaded source and configs frequently carry secrets; scrub the preview."""
    for pattern in _SECRET_PATTERNS:
        text = pattern.sub("[redacted]", text)
    return text


def extract_preview(path: Path, classification: Classification, char_cap: int = 20000) -> str:
    """Produce a Markdown preview for an accepted upload, by family."""
    family = classification.spec.family
    if classification.spec.handling == "header_only":
        return _header_summary(path, family)
    if classification.spec.handling != "extract":
        raise ExtractionError("extraction_not_supported")
    extractor = {
        Family.MARKDOWN: _markdown,
        Family.TEXT: _plain_text,
        Family.CODE: _source_code,
        Family.STRUCTURED: _structured,
        Family.TABULAR: _tabular,
        Family.NOTEBOOK: _notebook,
        Family.MARKUP: _html,
        Family.SPREADSHEET: _spreadsheet,
        Family.OFFICE: _office,
        Family.VECTOR: _svg,
    }.get(family, _plain_text)
    return extractor(path, char_cap)


def _decode(path: Path, char_cap: int) -> str:
    raw = path.read_bytes()[: char_cap * 4]
    for encoding in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            return raw.decode(encoding)[:char_cap]
        except UnicodeDecodeError:
            continue
    raise ExtractionError("extraction_encoding_failed")


def _markdown(path: Path, char_cap: int) -> str:
    return redact(_decode(path, char_cap))


def _plain_text(path: Path, char_cap: int) -> str:
    return f"```text\n{redact(_decode(path, char_cap))}\n```\n"


def _source_code(path: Path, char_cap: int) -> str:
    language = LANGUAGE_BY_EXTENSION.get(path.suffix.lower(), "text")
    body = redact(_decode(path, char_cap))
    lines = body.count("\n") + 1
    return (
        f"**Source file** `{path.name}` — {language}, {lines} lines\n\n```{language}\n{body}\n```\n"
    )


def _structured(path: Path, char_cap: int) -> str:
    body = redact(_decode(path, char_cap))
    language = LANGUAGE_BY_EXTENSION.get(path.suffix.lower(), "text")
    return f"```{language}\n{body}\n```\n"


def _tabular(path: Path, char_cap: int) -> str:
    suffix = path.suffix.lower()
    if suffix in {".jsonl", ".ndjson"}:
        records, keys, total = [], set(), 0
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            for index, line in enumerate(handle):
                total = index + 1
                if index < MAX_JSONL_SAMPLE_RECORDS and line.strip():
                    try:
                        record = json.loads(line)
                    except ValueError:
                        continue
                    records.append(record)
                    if isinstance(record, dict):
                        keys |= set(record)
        sample = json.dumps(records, indent=2)[:char_cap]
        key_list = ", ".join(sorted(str(key) for key in keys)[:40]) or "none"
        return (
            f"**JSONL dataset** `{path.name}` — {total} records; top-level keys: "
            f"{key_list}\n\n```json\n{redact(sample)}\n```\n"
        )

    delimiter = "\t" if suffix == ".tsv" else "|" if suffix == ".psv" else ","
    rows: list[list[str]] = []
    total = 0
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        for index, row in enumerate(csv.reader(handle, delimiter=delimiter)):
            total = index + 1
            if index <= MAX_TABULAR_SAMPLE_ROWS:
                rows.append(row)
    if not rows:
        return f"**Empty table** `{path.name}`\n"
    header, body = rows[0], rows[1:]
    table_lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    table_lines += [
        "| " + " | ".join(cell.replace("|", "\\|") for cell in row) + " |" for row in body
    ]
    rendered = redact("\n".join(table_lines))[:char_cap]
    return (
        f"**Table** `{path.name}` — {total} rows, {len(header)} columns "
        f"(showing first {len(body)})\n\n{rendered}\n"
    )


def _spreadsheet(path: Path, char_cap: int) -> str:
    """Read cell values only via openpyxl; VBA macros are never loaded or run."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExtractionError("extraction_missing_openpyxl") from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExtractionError("extraction_invalid_spreadsheet") from exc
    parts = [
        f"**Spreadsheet** `{path.name}` - {len(workbook.sheetnames)} sheet(s); "
        "values only, formulas shown as last-computed results, macros not executed.\n"
    ]
    for sheet_name in workbook.sheetnames[:5]:
        sheet = workbook[sheet_name]
        rows = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= 40:
                rows.append(["..."])
                break
            rows.append(["" if cell is None else str(cell) for cell in row[:20]])
        parts.append(f"### Sheet: {sheet_name} ({sheet.max_row} rows x {sheet.max_column} cols)")
        for row in rows:
            parts.append("| " + " | ".join(value.replace("|", "\\|") for value in row) + " |")
        parts.append("")
    if len(workbook.sheetnames) > 5:
        parts.append(f"({len(workbook.sheetnames) - 5} more sheet(s) not shown)")
    workbook.close()
    try:
        formula_book = load_workbook(path, read_only=True, data_only=False)
        formula_lines = []
        for sheet_name in formula_book.sheetnames[:5]:
            for row in formula_book[sheet_name].iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formula_lines.append(f"- {sheet_name}!{cell.coordinate}: `{value}`")
                        if len(formula_lines) >= 200:
                            break
                if len(formula_lines) >= 200:
                    break
            if len(formula_lines) >= 200:
                break
        formula_book.close()
        if formula_lines:
            parts.append("### Formulas (up to 200)")
            parts.extend(formula_lines)
    except Exception:
        pass
    return redact("\n".join(parts))[:char_cap]


def _office(path: Path, char_cap: int) -> str:
    """Extract text from OOXML containers via stdlib zip+XML; nothing executes."""
    import zipfile

    suffix = path.suffix.lower()
    texts = []
    try:
        with zipfile.ZipFile(path) as archive:
            if suffix == ".docx":
                members = ["word/document.xml"]
            else:
                members = sorted(
                    name
                    for name in archive.namelist()
                    if name.startswith("ppt/slides/slide") and name.endswith(".xml")
                )[:30]
            for member in members:
                root = ElementTree.fromstring(archive.read(member))
                runs = [node.text for node in root.iter() if node.tag.endswith("}t") and node.text]
                if runs:
                    label = member.rsplit("/", 1)[-1].replace(".xml", "")
                    texts.append((label, " ".join(runs)))
    except (zipfile.BadZipFile, ElementTree.ParseError, KeyError) as exc:
        raise ExtractionError("extraction_invalid_office") from exc
    kind = "Word document" if suffix == ".docx" else "Presentation"
    parts = [f"**{kind}** `{path.name}` - text extracted; embedded objects and macros untouched.\n"]
    for label, body in texts:
        if suffix == ".pptx":
            parts.append(f"### {label}")
        parts.append(body)
    return redact("\n\n".join(parts))[:char_cap]


def _notebook(path: Path, char_cap: int) -> str:
    """Notebooks are read as documents, never executed. Rich outputs are dropped."""
    try:
        document = json.loads(_decode(path, char_cap * 4))
    except ValueError as exc:
        raise ExtractionError("extraction_invalid_notebook") from exc
    metadata = document.get("metadata", {}) if isinstance(document, dict) else {}
    kernel = (metadata.get("kernelspec", {}) or {}).get("language") or "python"
    cells = document.get("cells", []) if isinstance(document, dict) else []
    parts = [f"**Notebook** `{path.name}` — {kernel}, {len(cells)} cells (outputs stripped)\n"]
    for index, cell in enumerate(cells, start=1):
        source = "".join(cell.get("source", []))
        if cell.get("cell_type") == "markdown":
            parts.append(source)
        elif cell.get("cell_type") == "code":
            parts.append(f"```{kernel}\n{source}\n```")
            text_output = _notebook_text_outputs(cell.get("outputs", []))
            if text_output:
                parts.append(
                    f"_Output {index} (text only, truncated):_\n```text\n{text_output[:2000]}\n```"
                )
    return redact("\n\n".join(parts))[:char_cap]


def _notebook_text_outputs(outputs: list) -> str:
    kept = []
    for output in outputs:
        if not isinstance(output, dict):
            continue
        data = output.get("data", {})
        if "text/plain" in data:
            kept.append("".join(data["text/plain"]))
        elif output.get("output_type") == "stream":
            kept.append("".join(output.get("text", [])))
        elif output.get("output_type") == "error":
            kept.append(f"{output.get('ename')}: {output.get('evalue')}")
    return "\n".join(kept)


_HTML_DROP_TAGS = {
    "script",
    "style",
    "iframe",
    "object",
    "embed",
    "applet",
    "link",
    "meta",
    "noscript",
    "template",
    "svg",
    "math",
    "form",
    "input",
    "button",
    "base",
}
_HTML_BLOCK_TAGS = {
    "p",
    "div",
    "section",
    "article",
    "br",
    "tr",
    "li",
    "blockquote",
    "pre",
    "table",
}
_SAFE_URL = re.compile(r"^(https?:|mailto:|#|/|\./|\.\./)", re.IGNORECASE)


class _HTMLToText(HTMLParser):
    """Allowlist tag stripper. No rendering, no script, no remote references."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.out: list[str] = []
        self._suppress_depth = 0
        self._pending_href: str | None = None

    def handle_starttag(self, tag, attrs):
        if tag in _HTML_DROP_TAGS:
            self._suppress_depth += 1
            return
        if self._suppress_depth:
            return
        if tag in {"h1", "h2", "h3", "h4", "h5", "h6"}:
            self.out.append("\n\n" + "#" * int(tag[1]) + " ")
        elif tag == "li":
            self.out.append("\n- ")
        elif tag in _HTML_BLOCK_TAGS:
            self.out.append("\n\n")
        elif tag == "a":
            href = (dict(attrs).get("href") or "").strip()
            if href and _SAFE_URL.match(href):
                self._pending_href = href

    def handle_endtag(self, tag):
        if tag in _HTML_DROP_TAGS and self._suppress_depth:
            self._suppress_depth -= 1
        elif tag == "a" and self._pending_href:
            self.out.append(f" <{self._pending_href}>")
            self._pending_href = None

    def handle_data(self, data):
        if not self._suppress_depth and data.strip():
            self.out.append(re.sub(r"[ \t]+", " ", data))


def _html(path: Path, char_cap: int) -> str:
    parser = _HTMLToText()
    parser.feed(_decode(path, char_cap * 2))
    parser.close()
    text = re.sub(r"\n{3,}", "\n\n", "".join(parser.out)).strip()[:char_cap]
    return (
        f"**HTML document** `{path.name}` — converted to text; scripts, styles, "
        f"frames, and forms removed.\n\n{redact(text)}\n"
    )


_SVG_FORBIDDEN_NODES = {
    "script",
    "foreignObject",
    "animate",
    "animateTransform",
    "set",
    "handler",
    "audio",
    "video",
    "iframe",
    "image",
}


def _svg(path: Path, char_cap: int) -> str:
    raw = path.read_bytes()
    if re.search(rb"<!DOCTYPE|<!ENTITY|SYSTEM\s|PUBLIC\s", raw[:4096], re.IGNORECASE):
        raise ExtractionError("extraction_unsafe_svg")
    try:
        root = ElementTree.fromstring(raw.decode("utf-8", errors="strict"))
    except (ElementTree.ParseError, UnicodeDecodeError) as exc:
        raise ExtractionError("extraction_unsafe_svg") from exc
    unsafe = 0
    for element in root.iter():
        if _svg_local(element.tag) in _SVG_FORBIDDEN_NODES:
            unsafe += 1
        for attribute, value in element.attrib.items():
            local = _svg_local(attribute).lower()
            external_ref = local in {"href", "src"} and not value.startswith("#")
            if local.startswith("on") or "javascript:" in value.lower() or external_ref:
                unsafe += 1
    labels = [
        (element.text or "").strip()
        for element in root.iter()
        if _svg_local(element.tag) in {"title", "desc", "text"} and (element.text or "").strip()
    ]
    text = "; ".join(labels[:50]) or "none"
    return (
        f"**SVG asset** `{path.name}` — inspected without rendering; "
        f"{unsafe} unsafe node(s) or attribute(s) present.\n\n"
        f"Text content: {redact(text)[:char_cap]}\n"
    )


def _svg_local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _header_summary(path: Path, family: Family) -> str:
    """Bounded metadata read for ML artifacts. Tensors are never materialized."""
    size = path.stat().st_size
    if family == Family.ML_TENSORS and path.suffix.lower() == ".safetensors":
        return _safetensors_header(path, size)
    if family == Family.ML_TENSORS:
        return _gguf_header(path, size)
    if family == Family.ML_ARRAY and path.suffix.lower() == ".npy":
        return _npy_header(path, size)
    return (
        f"**Binary artifact** `{path.name}` — {size} bytes, registered without "
        f"parsing; structural metadata only.\n"
    )


def _safetensors_header(path: Path, size: int) -> str:
    with path.open("rb") as handle:
        prefix = handle.read(8)
        if len(prefix) < 8:
            raise ExtractionError("extraction_invalid_header")
        (header_length,) = struct.unpack("<Q", prefix)
        if header_length > 16 * 1024 * 1024:
            raise ExtractionError("extraction_invalid_header")
        try:
            header = json.loads(handle.read(header_length).decode("utf-8"))
        except (ValueError, UnicodeDecodeError) as exc:
            raise ExtractionError("extraction_invalid_header") from exc
    tensors = {name: spec for name, spec in header.items() if name != "__metadata__"}
    dtypes = sorted({str(spec.get("dtype")) for spec in tensors.values() if isinstance(spec, dict)})
    return (
        f"**Safetensors weights** `{path.name}` — {size} bytes, "
        f"{len(tensors)} tensors, dtypes: {', '.join(dtypes) or 'unknown'}. "
        f"Tensors were not loaded.\n"
    )


def _gguf_header(path: Path, size: int) -> str:
    with path.open("rb") as handle:
        magic = handle.read(4)
        if magic != b"GGUF":
            raise ExtractionError("extraction_invalid_header")
        version_bytes = handle.read(4)
        tensor_count_bytes = handle.read(8)
        if len(version_bytes) < 4 or len(tensor_count_bytes) < 8:
            raise ExtractionError("extraction_invalid_header")
        (version,) = struct.unpack("<I", version_bytes)
        (tensor_count,) = struct.unpack("<Q", tensor_count_bytes)
    return (
        f"**GGUF weights** `{path.name}` — {size} bytes, format version {version}, "
        f"{tensor_count} tensors. Tensors were not loaded.\n"
    )


def _npy_header(path: Path, size: int) -> str:
    with path.open("rb") as handle:
        magic = handle.read(6)
        if magic != b"\x93NUMPY":
            raise ExtractionError("extraction_invalid_header")
        handle.read(2)
        (header_length,) = struct.unpack("<H", handle.read(2))
        header = handle.read(min(header_length, 64 * 1024)).decode("latin-1", errors="replace")
    shape_match = re.search(r"'shape':\s*\(([^)]*)\)", header)
    dtype_match = re.search(r"'descr':\s*'([^']*)'", header)
    shape = shape_match.group(1).strip().rstrip(",") if shape_match else "unknown"
    dtype = dtype_match.group(1) if dtype_match else "unknown"
    return (
        f"**NumPy array** `{path.name}` — {size} bytes, shape ({shape}), "
        f"dtype {dtype}. The array was not materialized.\n"
    )
